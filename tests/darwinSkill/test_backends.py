from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

import openpyxl

from darwinSkill.backends import (
    ALFWorldEpisodeTargetBackend,
    BackendRouter,
    BackendRuntimeConfig,
    SpreadsheetReactTargetBackend,
    build_alfworld_router,
    build_spreadsheetbench_router,
    default_routing_for_family,
)
from darwinSkill.contracts import MetricResult, SkillFeedback, SkillSample


class TargetStub:
    def predict(self, skill_text: str, sample: SkillSample) -> str:
        return f"target:{sample.prompt}"


class OptimizerStub:
    def improve_skill(self, skill_text: str, feedback: list[SkillFeedback]) -> str:
        return f"{skill_text}|optimizer:{len(feedback)}"


class ScriptedSpreadsheetBackend:
    def __init__(self) -> None:
        self.turn = 0

    def respond(self, *, messages, tools, tool_choice="auto"):  # type: ignore[no-untyped-def]
        _ = (messages, tools, tool_choice)
        self.turn += 1
        if self.turn == 1:
            return {
                "content": "Write and run the workbook patch.",
                "tool_calls": [
                    {
                        "id": "write-1",
                        "name": "write_file",
                        "arguments": {
                            "path": "solution.py",
                            "content": (
                                "import openpyxl\n"
                                "wb = openpyxl.load_workbook(INPUT_PATH)\n"
                                "ws = wb['Sheet1']\n"
                                "ws['A1'] = 77\n"
                                "wb.save(OUTPUT_PATH)\n"
                            ),
                        },
                    },
                    {
                        "id": "bash-1",
                        "name": "bash",
                        "arguments": {"cmd": "python solution.py"},
                    },
                ],
            }
        return {"content": "Finished."}


class ScriptedALFWorldBackend:
    def __init__(self) -> None:
        self.turn = 0

    def respond(self, *, system: str, user: str, messages=None):  # type: ignore[no-untyped-def]
        _ = (system, user, messages)
        self.turn += 1
        if self.turn == 1:
            return "<think>inspect</think><action>look</action>"
        return "<think>finish</think><action>put apple in sink</action>"


class FakeALFWorldEnvironment:
    def __init__(self) -> None:
        self.turn = 0

    def reset(self):  # type: ignore[no-untyped-def]
        self.turn = 0
        return {
            "observation": "You are in a kitchen. Your task is to: put apple in sink.",
            "gamefile": "/tmp/valid_seen/pick_and_place/task.json",
            "task_type": "pick_and_place",
            "task_description": "put apple in sink",
        }

    def step(self, action):  # type: ignore[no-untyped-def]
        self.turn += 1
        if self.turn == 1:
            return {"observation": f"after {action}", "reward": 0.0, "done": False, "won": False}
        return {"observation": f"after {action}", "reward": 1.0, "done": True, "won": True}


class BackendsTest(unittest.TestCase):
    def test_backend_router_separates_target_and_optimizer_roles(self) -> None:
        router = BackendRouter(target_backend=TargetStub(), optimizer_backend=OptimizerStub())
        sample = SkillSample(prompt="Capital of France?", expected_answer="Paris")
        feedback = [
            SkillFeedback(
                sample=sample,
                prediction="unknown",
                metric=MetricResult(score=0.0, passed=False),
            )
        ]

        self.assertEqual(router.predict("", sample), "target:Capital of France?")
        self.assertEqual(router.improve_skill("", feedback), "|optimizer:1")

    def test_backend_runtime_config_rejects_unknown_family(self) -> None:
        with self.assertRaises(ValueError):
            BackendRuntimeConfig(family="unknown")

    def test_default_routing_family_mapping(self) -> None:
        routing = default_routing_for_family("codex_exec")
        self.assertEqual(routing.target.family, "codex_exec")
        self.assertEqual(routing.optimizer.family, "openai_chat")

    def test_spreadsheet_react_target_backend_wraps_chat_backend(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-1"
            task_dir.mkdir(parents=True)
            for path in (task_dir / "initial.xlsx", task_dir / "golden.xlsx"):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["A1"] = 1
                workbook.save(path)
                workbook.close()
            sample = SkillSample(
                prompt="Set A1 to 77",
                expected_answer="",
                metadata={
                    "instruction_type": "cell update",
                    "answer_position": "Sheet1!A1",
                    "spreadsheet_path": "spreadsheet/task-1",
                    "data_root": str(root),
                },
            )
            backend = SpreadsheetReactTargetBackend(backend=ScriptedSpreadsheetBackend())
            prediction = backend.predict("", sample)
            payload = json.loads(prediction)
            self.assertIn("conversation", payload)
            self.assertIn("solution.py", payload["artifacts"])

    def test_alfworld_episode_target_backend_wraps_chat_backend(self) -> None:
        sample = SkillSample(
            prompt="put apple in sink",
            expected_answer="success",
            metadata={
                "gamefile": "/tmp/valid_seen/pick_and_place/task.json",
                "task_type": "pick_and_place",
                "task_description": "put apple in sink",
            },
        )
        backend = ALFWorldEpisodeTargetBackend(
            backend=ScriptedALFWorldBackend(),
            environment_factory=lambda current_sample: FakeALFWorldEnvironment(),
        )
        prediction = backend.predict("", sample)
        payload = json.loads(prediction)
        self.assertEqual(payload["hard"], 1)
        self.assertEqual(len(payload["conversation"]), 2)

    def test_build_spreadsheetbench_router_wraps_target_role(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-1"
            task_dir.mkdir(parents=True)
            for path in (task_dir / "initial.xlsx", task_dir / "golden.xlsx"):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["A1"] = 1
                workbook.save(path)
                workbook.close()
            router = build_spreadsheetbench_router(
                target_backend=ScriptedSpreadsheetBackend(),
                optimizer_backend=OptimizerStub(),
            )
            sample = SkillSample(
                prompt="Set A1 to 77",
                expected_answer="",
                metadata={
                    "instruction_type": "cell update",
                    "answer_position": "Sheet1!A1",
                    "spreadsheet_path": "spreadsheet/task-1",
                    "data_root": str(root),
                },
            )
            prediction = router.predict("", sample)
            payload = json.loads(prediction)
            self.assertIn("conversation", payload)

    def test_build_alfworld_router_wraps_target_role(self) -> None:
        router = build_alfworld_router(
            target_backend=ScriptedALFWorldBackend(),
            optimizer_backend=OptimizerStub(),
            environment_factory=lambda sample: FakeALFWorldEnvironment(),
        )
        sample = SkillSample(
            prompt="put apple in sink",
            expected_answer="success",
            metadata={"gamefile": "/tmp/valid_seen/pick_and_place/task.json"},
        )
        prediction = router.predict("", sample)
        payload = json.loads(prediction)
        self.assertEqual(payload["hard"], 1)


if __name__ == "__main__":
    unittest.main()

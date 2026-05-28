from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

import openpyxl

from darwinSkill.alfworld_env import ALFWorldEvaluator, load_alfworld_dataset, run_alfworld_episode
from darwinSkill.contracts import SkillFeedback, TrainingConfig
from darwinSkill.livemathematician_env import LiveMathematicianEvaluator, load_livemathematician_dataset
from darwinSkill.native import run_reference_adapter
from darwinSkill.reference_adapters import ALFWorldAdapter, LiveMathematicianBenchAdapter, SpreadsheetBenchAdapter
from darwinSkill.spreadsheetbench_env import (
    SpreadsheetBenchEvaluator,
    load_spreadsheetbench_dataset,
    run_spreadsheet_react_session,
)


class PackBBackend:
    def predict(self, skill_text, sample):  # type: ignore[no-untyped-def]
        metadata = sample.metadata
        if "correct_choice" in metadata:
            return f"<answer>{metadata['correct_choice']['label']}</answer>"
        if metadata.get("pred_path"):
            return f"<answer>{metadata['pred_path']}</answer>"
        return "<answer>success</answer>"

    def improve_skill(self, skill_text: str, feedback: list[SkillFeedback]) -> str:
        return skill_text


class ScriptedSpreadsheetReactBackend:
    def __init__(self) -> None:
        self.turn = 0

    def respond(self, *, messages, tools, tool_choice="auto"):  # type: ignore[no-untyped-def]
        _ = (messages, tools, tool_choice)
        self.turn += 1
        if self.turn == 1:
            return {
                "content": "I will write and execute solution.py.",
                "tool_calls": [
                    {
                        "id": "write-1",
                        "name": "write_file",
                        "arguments": {
                            "path": "solution.py",
                            "content": (
                                "import openpyxl\n"
                                "wb = openpyxl.load_workbook(INPUT_PATH)\n"
                                "ws = wb['Sheet1']\n"
                                "ws['G7'] = 2024\n"
                                "wb.save(OUTPUT_PATH)\n"
                            ),
                        },
                    },
                    {
                        "id": "bash-1",
                        "name": "bash",
                        "arguments": {"cmd": "python solution.py"},
                    },
                ],
            }
        return {"content": "Finished."}


class SpreadsheetReactPredictionBackend:
    def improve_skill(self, skill_text: str, feedback: list[SkillFeedback]) -> str:
        return skill_text

    def predict(self, skill_text, sample):  # type: ignore[no-untyped-def]
        backend = ScriptedSpreadsheetReactBackend()
        session = run_spreadsheet_react_session(
            backend=backend,
            sample=sample,
            skill_content=skill_text,
        )
        return str(session["prediction"])


class ScriptedALFWorldBackend:
    def __init__(self) -> None:
        self.turn = 0

    def respond(self, *, system: str, user: str, messages=None):  # type: ignore[no-untyped-def]
        _ = (system, user, messages)
        self.turn += 1
        if self.turn == 1:
            return "<think>Need to inspect the room first.</think><action>look</action>"
        return "<think>I can finish the task now.</think><action>put apple in sink</action>"


class FakeALFWorldEnvironment:
    def __init__(self) -> None:
        self.turn = 0

    def reset(self):  # type: ignore[no-untyped-def]
        self.turn = 0
        return {
            "observation": "You are in a kitchen. Your task is to: put apple in sink.",
            "gamefile": "/tmp/valid_seen/pick_and_place/task.json",
            "task_type": "pick_and_place",
            "task_description": "put apple in sink",
        }

    def step(self, action):  # type: ignore[no-untyped-def]
        self.turn += 1
        if self.turn == 1:
            return {
                "observation": f"Observed after {action}: you see an apple and a sink.",
                "reward": 0.0,
                "done": False,
                "won": False,
            }
        return {
            "observation": f"Observed after {action}: task completed.",
            "reward": 1.0,
            "done": True,
            "won": True,
        }


class ALFWorldRuntimePredictionBackend:
    def improve_skill(self, skill_text: str, feedback: list[SkillFeedback]) -> str:
        return skill_text

    def predict(self, skill_text, sample):  # type: ignore[no-untyped-def]
        episode = run_alfworld_episode(
            backend=ScriptedALFWorldBackend(),
            environment=FakeALFWorldEnvironment(),
            skill_content=skill_text,
        )
        return str(episode["prediction"])


class ReferencePackBEnvTest(unittest.TestCase):
    def test_livemathematician_loader_evaluator_and_native_flow(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            for split, month, label in (("train", "2024-01", "B"), ("val", "2024-02", "C"), ("test", "2024-03", "A")):
                split_dir = root / split
                split_dir.mkdir(parents=True)
                payload = [
                    {
                        "month": month,
                        "no": 1,
                        "theorem_type": ["algebra"],
                        "mcq": {
                            "question": f"Question {split}?",
                            "choices": [{"label": "A", "text": "One"}, {"label": "B", "text": "Two"}, {"label": "C", "text": "Three"}],
                            "correct_choice": {"label": label, "text": {"A": "One", "B": "Two", "C": "Three"}[label]},
                        },
                    }
                ]
                (split_dir / "qa_mock_final.json").write_text(json.dumps(payload), encoding="utf-8")

            train_records, eval_records = load_livemathematician_dataset(root)
            self.assertEqual(train_records[0]["correct_choice"]["label"], "B")
            self.assertEqual(eval_records[0]["correct_choice"]["label"], "C")

            adapter = LiveMathematicianBenchAdapter.from_path(str(root))
            metric = LiveMathematicianEvaluator().evaluate("<answer>B</answer>", adapter.train_samples[0])
            self.assertTrue(metric.passed)

            artifacts = run_reference_adapter(
                backend=PackBBackend(),
                adapter=adapter,
                config=TrainingConfig(
                    num_epochs=1,
                    batch_size=1,
                    edit_budget=1,
                    output_root=root / "outputs",
                    run_name="livemath-pack-b",
                ),
            )
            self.assertEqual(artifacts.mean_score, 1.0)

    def test_spreadsheet_loader_evaluator_and_native_flow(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            gold_path = root / "gold.xlsx"
            pred_path = root / "pred.xlsx"
            for path in (gold_path, pred_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["A1"] = 42
                workbook.save(path)
                workbook.close()

            for split in ("train", "val", "test"):
                split_dir = root / split
                split_dir.mkdir(parents=True)
                payload = [
                    {
                        "id": split,
                        "instruction": f"Fill value for {split}",
                        "instruction_type": "cell update",
                        "answer_position": "Sheet1!A1",
                        "gold_path": str(gold_path),
                        "pred_path": str(pred_path),
                    }
                ]
                (split_dir / "items.json").write_text(json.dumps(payload), encoding="utf-8")

            train_records, eval_records = load_spreadsheetbench_dataset(root)
            self.assertEqual(train_records[0]["task_type"], "cell_level")
            self.assertEqual(eval_records[0]["answer_position"], "Sheet1!A1")

            adapter = SpreadsheetBenchAdapter.from_path(str(root))
            metric = SpreadsheetBenchEvaluator().evaluate(f"<answer>{pred_path}</answer>", adapter.train_samples[0])
            self.assertTrue(metric.passed)

            artifacts = run_reference_adapter(
                backend=PackBBackend(),
                adapter=adapter,
                config=TrainingConfig(
                    num_epochs=1,
                    batch_size=1,
                    edit_budget=1,
                    output_root=root / "outputs",
                    run_name="spreadsheet-pack-b",
                ),
            )
            self.assertEqual(artifacts.mean_score, 1.0)

    def test_spreadsheet_evaluator_can_execute_generated_code(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-1"
            task_dir.mkdir(parents=True)
            input_path = task_dir / "1_task_input.xlsx"
            gold_path = task_dir / "1_task_answer.xlsx"
            for path in (input_path, gold_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["A1"] = 1
                workbook.save(path)
                workbook.close()
            golden = openpyxl.load_workbook(gold_path)
            golden["Sheet1"]["A1"] = 42
            golden.save(gold_path)
            golden.close()

            adapter = SpreadsheetBenchAdapter.from_records(
                [
                    {
                        "id": "task-1",
                        "instruction": "Set A1 to 42",
                        "instruction_type": "cell update",
                        "answer_position": "Sheet1!A1",
                        "spreadsheet_path": "spreadsheet/task-1",
                        "data_root": str(root),
                    }
                ]
            )
            metric = SpreadsheetBenchEvaluator().evaluate(
                """```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["Sheet1"]
ws["A1"] = 42
wb.save(OUTPUT_PATH)
```""",
                adapter.train_samples[0],
            )
            self.assertTrue(metric.passed)
            self.assertEqual(metric.details["hard"], 1)

    def test_spreadsheet_evaluator_accepts_json_artifact_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-2"
            task_dir.mkdir(parents=True)
            input_path = task_dir / "1_task_init.xlsx"
            gold_path = task_dir / "1_task_golden.xlsx"
            for path in (input_path, gold_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["B2"] = 5
                workbook.save(path)
                workbook.close()
            golden = openpyxl.load_workbook(gold_path)
            golden["Sheet1"]["B2"] = 99
            golden.save(gold_path)
            golden.close()

            adapter = SpreadsheetBenchAdapter.from_records(
                [
                    {
                        "id": "task-2",
                        "instruction": "Set B2 to 99",
                        "instruction_type": "cell update",
                        "answer_position": "Sheet1!B2",
                        "spreadsheet_path": "spreadsheet/task-2",
                        "data_root": str(root),
                    }
                ]
            )
            prediction = json.dumps(
                {
                    "artifacts": {
                        "solution.py": (
                            "import openpyxl\n"
                            "wb = openpyxl.load_workbook(INPUT_PATH)\n"
                            "ws = wb['Sheet1']\n"
                            "ws['B2'] = 99\n"
                            "wb.save(OUTPUT_PATH)\n"
                        )
                    }
                }
            )
            metric = SpreadsheetBenchEvaluator().evaluate(prediction, adapter.train_samples[0])
            self.assertTrue(metric.passed)
            self.assertEqual(metric.details["mode"], "generated_code_bundle")

    def test_spreadsheet_evaluator_accepts_workspace_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-3"
            task_dir.mkdir(parents=True)
            input_path = task_dir / "initial.xlsx"
            gold_path = task_dir / "golden.xlsx"
            for path in (input_path, gold_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["C3"] = 7
                workbook.save(path)
                workbook.close()
            golden = openpyxl.load_workbook(gold_path)
            golden["Sheet1"]["C3"] = 123
            golden.save(gold_path)
            golden.close()

            adapter = SpreadsheetBenchAdapter.from_records(
                [
                    {
                        "id": "task-3",
                        "instruction": "Set C3 to 123",
                        "instruction_type": "cell update",
                        "answer_position": "Sheet1!C3",
                        "spreadsheet_path": "spreadsheet/task-3",
                        "data_root": str(root),
                    }
                ]
            )
            prediction = json.dumps(
                {
                    "files": {
                        "solution.py": (
                            "import openpyxl\n"
                            "wb = openpyxl.load_workbook(INPUT_PATH)\n"
                            "ws = wb['Sheet1']\n"
                            "ws['C3'] = 123\n"
                            "wb.save(OUTPUT_PATH)\n"
                        )
                    },
                    "commands": ["python solution.py"],
                }
            )
            metric = SpreadsheetBenchEvaluator().evaluate(prediction, adapter.train_samples[0])
            self.assertTrue(metric.passed)
            self.assertEqual(metric.details["mode"], "workspace_bundle")

    def test_spreadsheet_evaluator_accepts_tool_call_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-4"
            task_dir.mkdir(parents=True)
            input_path = task_dir / "initial.xlsx"
            gold_path = task_dir / "golden.xlsx"
            for path in (input_path, gold_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["D4"] = 11
                workbook.save(path)
                workbook.close()
            golden = openpyxl.load_workbook(gold_path)
            golden["Sheet1"]["D4"] = 555
            golden.save(gold_path)
            golden.close()

            adapter = SpreadsheetBenchAdapter.from_records(
                [
                    {
                        "id": "task-4",
                        "instruction": "Set D4 to 555",
                        "instruction_type": "cell update",
                        "answer_position": "Sheet1!D4",
                        "spreadsheet_path": "spreadsheet/task-4",
                        "data_root": str(root),
                    }
                ]
            )
            prediction = json.dumps(
                {
                    "tool_calls": [
                        {
                            "function": {
                                "name": "write_file",
                                "arguments": {
                                    "path": "solution.py",
                                    "content": (
                                        "import openpyxl\n"
                                        "wb = openpyxl.load_workbook(INPUT_PATH)\n"
                                        "ws = wb['Sheet1']\n"
                                        "ws['D4'] = 555\n"
                                        "wb.save(OUTPUT_PATH)\n"
                                    ),
                                },
                            }
                        },
                        {
                            "function": {
                                "name": "bash",
                                "arguments": {"cmd": "python solution.py"},
                            }
                        },
                    ]
                }
            )
            metric = SpreadsheetBenchEvaluator().evaluate(prediction, adapter.train_samples[0])
            self.assertTrue(metric.passed)
            self.assertEqual(metric.details["mode"], "tool_call_bundle")

    def test_spreadsheet_evaluator_accepts_react_transcript_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-5"
            task_dir.mkdir(parents=True)
            input_path = task_dir / "initial.xlsx"
            gold_path = task_dir / "golden.xlsx"
            for path in (input_path, gold_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["E5"] = 21
                workbook.save(path)
                workbook.close()
            golden = openpyxl.load_workbook(gold_path)
            golden["Sheet1"]["E5"] = 777
            golden.save(gold_path)
            golden.close()

            adapter = SpreadsheetBenchAdapter.from_records(
                [
                    {
                        "id": "task-5",
                        "instruction": "Set E5 to 777",
                        "instruction_type": "cell update",
                        "answer_position": "Sheet1!E5",
                        "spreadsheet_path": "spreadsheet/task-5",
                        "data_root": str(root),
                    }
                ]
            )
            prediction = json.dumps(
                {
                    "transcript": [
                        {
                            "role": "assistant",
                            "content": "I will write a solution and run it.",
                            "tool_calls": [
                                {
                                    "function": {
                                        "name": "write_file",
                                        "arguments": {
                                            "path": "solution.py",
                                            "content": (
                                                "import openpyxl\n"
                                                "wb = openpyxl.load_workbook(INPUT_PATH)\n"
                                                "ws = wb['Sheet1']\n"
                                                "ws['E5'] = 777\n"
                                                "wb.save(OUTPUT_PATH)\n"
                                            ),
                                        },
                                    }
                                },
                                {
                                    "function": {
                                        "name": "bash",
                                        "arguments": {"cmd": "python solution.py"},
                                    }
                                },
                            ],
                        },
                        {"type": "message", "content": "Done."},
                    ]
                }
            )
            metric = SpreadsheetBenchEvaluator().evaluate(prediction, adapter.train_samples[0])
            self.assertTrue(metric.passed)
            self.assertEqual(metric.details["mode"], "react_transcript_bundle")

    def test_spreadsheet_evaluator_accepts_upstream_conversation_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-6"
            task_dir.mkdir(parents=True)
            input_path = task_dir / "initial.xlsx"
            gold_path = task_dir / "golden.xlsx"
            for path in (input_path, gold_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["F6"] = 31
                workbook.save(path)
                workbook.close()
            golden = openpyxl.load_workbook(gold_path)
            golden["Sheet1"]["F6"] = 909
            golden.save(gold_path)
            golden.close()

            adapter = SpreadsheetBenchAdapter.from_records(
                [
                    {
                        "id": "task-6",
                        "instruction": "Set F6 to 909",
                        "instruction_type": "cell update",
                        "answer_position": "Sheet1!F6",
                        "spreadsheet_path": "spreadsheet/task-6",
                        "data_root": str(root),
                    }
                ]
            )
            prediction = json.dumps(
                {
                    "conversation": [
                        {"role": "assistant", "content": "I will write solution.py and run it."},
                        {
                            "type": "tool_call",
                            "cmd": "[write_file] solution.py",
                            "obs": "File written: /tmp/solution.py (99 chars)",
                        },
                        {
                            "type": "tool_call",
                            "cmd": "python solution.py",
                            "obs": "Execution completed.",
                        },
                    ],
                    "artifacts": {
                        "solution.py": (
                            "import openpyxl\n"
                            "wb = openpyxl.load_workbook(INPUT_PATH)\n"
                            "ws = wb['Sheet1']\n"
                            "ws['F6'] = 909\n"
                            "wb.save(OUTPUT_PATH)\n"
                        )
                    },
                }
            )
            metric = SpreadsheetBenchEvaluator().evaluate(prediction, adapter.train_samples[0])
            self.assertTrue(metric.passed)
            self.assertEqual(metric.details["mode"], "react_conversation_bundle")

    def test_spreadsheet_react_session_builds_prediction_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-7"
            task_dir.mkdir(parents=True)
            input_path = task_dir / "initial.xlsx"
            gold_path = task_dir / "golden.xlsx"
            for path in (input_path, gold_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["G7"] = 12
                workbook.save(path)
                workbook.close()
            golden = openpyxl.load_workbook(gold_path)
            golden["Sheet1"]["G7"] = 2024
            golden.save(gold_path)
            golden.close()

            adapter = SpreadsheetBenchAdapter.from_records(
                [
                    {
                        "id": "task-7",
                        "instruction": "Set G7 to 2024",
                        "instruction_type": "cell update",
                        "answer_position": "Sheet1!G7",
                        "spreadsheet_path": "spreadsheet/task-7",
                        "data_root": str(root),
                    }
                ]
            )
            session = run_spreadsheet_react_session(
                backend=ScriptedSpreadsheetReactBackend(),
                sample=adapter.train_samples[0],
                skill_content="Always verify the workbook after saving.",
            )
            self.assertIn("solution.py", session["artifacts"])
            self.assertGreaterEqual(int(session["n_turns"]), 2)

            metric = SpreadsheetBenchEvaluator().evaluate(str(session["prediction"]), adapter.train_samples[0])
            self.assertTrue(metric.passed)
            self.assertEqual(metric.details["mode"], "react_conversation_bundle")

    def test_spreadsheet_react_prediction_backend_runs_with_native_adapter_flow(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task_dir = root / "spreadsheet" / "task-8"
            task_dir.mkdir(parents=True)
            input_path = task_dir / "initial.xlsx"
            gold_path = task_dir / "golden.xlsx"
            for path in (input_path, gold_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet["G7"] = 12
                workbook.save(path)
                workbook.close()
            golden = openpyxl.load_workbook(gold_path)
            golden["Sheet1"]["G7"] = 2024
            golden.save(gold_path)
            golden.close()

            adapter = SpreadsheetBenchAdapter.from_records(
                [
                    {
                        "id": "task-8",
                        "instruction": "Set G7 to 2024",
                        "instruction_type": "cell update",
                        "answer_position": "Sheet1!G7",
                        "spreadsheet_path": "spreadsheet/task-8",
                        "data_root": str(root),
                    }
                ]
            )
            artifacts = run_reference_adapter(
                backend=SpreadsheetReactPredictionBackend(),
                adapter=adapter,
                config=TrainingConfig(
                    num_epochs=1,
                    batch_size=1,
                    edit_budget=1,
                    output_root=root / "outputs",
                    run_name="spreadsheet-react-pack-b",
                ),
            )
            self.assertEqual(artifacts.mean_score, 1.0)

    def test_alfworld_loader_evaluator_and_native_flow(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            records_by_split = {
                "train": [{"id": "train-1", "gamefile": "/tmp/train/task.json", "task_description": "Put apple in sink", "task_type": "pick_and_place"}],
                "val": [{"id": "val-1", "gamefile": "/tmp/valid_seen/task.json", "task_description": "Heat soup", "task_type": "heat"}],
                "test": [{"id": "test-1", "gamefile": "/tmp/valid_unseen/task.json", "task_description": "Cool drink", "task_type": "cool"}],
            }
            for split, payload in records_by_split.items():
                split_dir = root / split
                split_dir.mkdir(parents=True)
                (split_dir / "items.json").write_text(json.dumps(payload), encoding="utf-8")

            train_records, eval_records = load_alfworld_dataset(root)
            self.assertEqual(train_records[0]["eval_dataset"], "train")
            self.assertEqual(eval_records[0]["eval_dataset"], "eval_in_distribution")

            adapter = ALFWorldAdapter.from_path(str(root))
            metric = ALFWorldEvaluator().evaluate("<answer>success</answer>", adapter.train_samples[0])
            self.assertTrue(metric.passed)

            artifacts = run_reference_adapter(
                backend=PackBBackend(),
                adapter=adapter,
                config=TrainingConfig(
                    num_epochs=1,
                    batch_size=1,
                    edit_budget=1,
                    output_root=root / "outputs",
                    run_name="alfworld-pack-b",
                ),
            )
            self.assertEqual(artifacts.mean_score, 1.0)

    def test_alfworld_runtime_episode_bundle(self) -> None:
        sample = ALFWorldAdapter.from_records(
            [
                {
                    "id": "episode-1",
                    "gamefile": "/tmp/valid_seen/pick_and_place/task.json",
                    "task_description": "put apple in sink",
                    "task_type": "pick_and_place",
                }
            ]
        ).train_samples[0]
        episode = run_alfworld_episode(
            backend=ScriptedALFWorldBackend(),
            environment=FakeALFWorldEnvironment(),
            skill_content="Inspect before acting.",
        )
        metric = ALFWorldEvaluator().evaluate(str(episode["prediction"]), sample)
        self.assertTrue(metric.passed)
        self.assertEqual(metric.details["mode"], "runtime_bundle")
        self.assertEqual(metric.details["n_turns"], 2)
        self.assertEqual(len(metric.details["conversation"]), 2)

    def test_alfworld_runtime_prediction_backend_runs_with_native_adapter_flow(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            adapter = ALFWorldAdapter.from_records(
                [
                    {
                        "id": "runtime-alfworld",
                        "gamefile": "/tmp/valid_seen/pick_and_place/task.json",
                        "task_description": "put apple in sink",
                        "task_type": "pick_and_place",
                    }
                ]
            )
            artifacts = run_reference_adapter(
                backend=ALFWorldRuntimePredictionBackend(),
                adapter=adapter,
                config=TrainingConfig(
                    num_epochs=1,
                    batch_size=1,
                    edit_budget=1,
                    output_root=root / "outputs",
                    run_name="alfworld-runtime-pack-b",
                ),
            )
            self.assertEqual(artifacts.mean_score, 1.0)


if __name__ == "__main__":
    unittest.main()

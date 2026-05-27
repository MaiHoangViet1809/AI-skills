from __future__ import annotations

import datetime
import json
import os
import re
import subprocess
import sys
import tempfile
import textwrap
from pathlib import Path
from typing import Any

import openpyxl

from darwinSkill.contracts import MetricResult, SkillEvaluator, SkillSample
from darwinSkill.reference_assets import is_split_dir


def _load_json_or_jsonl(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            payload = payload.get("data") or list(payload.values())
        if not isinstance(payload, list):
            raise ValueError(f"Expected list payload in {path}")
        return [dict(item) for item in payload]
    items: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            items.append(json.loads(line))
    return items


def _resolve_data_file(path: Path) -> Path:
    if path.is_file():
        return path
    candidates = sorted(path.glob("*.json")) + sorted(path.glob("*.jsonl"))
    if len(candidates) != 1:
        raise ValueError(f"SpreadsheetBench expects exactly one JSON/JSONL file under {path}")
    return candidates[0]


def _infer_task_type(instruction_type: str) -> str:
    lowered = instruction_type.lower()
    if "cell" in lowered:
        return "cell_level"
    if "sheet" in lowered:
        return "sheet_level"
    return "other"


PATH_ASSIGN_RE = re.compile(r"^\s*(INPUT_PATH|OUTPUT_PATH)\s*=\s*.+$", re.MULTILINE)
RUNNER_TEMPLATE = textwrap.dedent(
    """
    import sys
    import traceback

    INPUT_PATH = {input_path!r}
    OUTPUT_PATH = {output_path!r}

    try:
    {user_code_indented}
    except Exception:
        traceback.print_exc()
        sys.exit(2)
    """
)


def normalize_spreadsheetbench_record(record: dict[str, Any]) -> dict[str, Any]:
    instruction_type = str(record.get("instruction_type") or "").strip()
    answer_sheet = str(record.get("answer_sheet") or "").strip()
    answer_position = str(record.get("answer_position") or "").strip()
    if answer_sheet and answer_position and "!" not in answer_position:
        answer_position = f"{answer_sheet}!{answer_position}"
    return {
        "id": str(record.get("id") or "").strip(),
        "instruction": str(record.get("instruction") or "").strip(),
        "instruction_type": instruction_type,
        "task_type": str(record.get("task_type") or _infer_task_type(instruction_type)),
        "spreadsheet_path": str(record.get("spreadsheet_path") or "").strip(),
        "answer_position": answer_position,
        "answer_sheet": answer_sheet,
        "gold_path": str(record.get("gold_path") or "").strip(),
        "pred_path": str(record.get("pred_path") or "").strip(),
        "expected_answer": str(record.get("expected_answer") or "").strip(),
        "data_root": str(record.get("data_root") or "").strip(),
        **{
            key: value
            for key, value in record.items()
            if key
            not in {
                "id",
                "instruction",
                "instruction_type",
                "task_type",
                "spreadsheet_path",
                "answer_position",
                "answer_sheet",
                "gold_path",
                "pred_path",
                "expected_answer",
                "data_root",
            }
        },
    }


def load_spreadsheetbench_records(path: Path | str) -> list[dict[str, Any]]:
    resolved = _resolve_data_file(Path(path))
    return [normalize_spreadsheetbench_record(item) for item in _load_json_or_jsonl(resolved)]


def load_spreadsheetbench_dataset(path: Path | str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    root = Path(path)
    if is_split_dir(root):
        train_records = load_spreadsheetbench_records(root / "train")
        eval_records = load_spreadsheetbench_records(root / "val")
        if not eval_records:
            eval_records = load_spreadsheetbench_records(root / "test")
        return train_records, eval_records or list(train_records)
    records = load_spreadsheetbench_records(root)
    return records, list(records)


def build_spreadsheetbench_samples(records: list[dict[str, Any]]) -> list[SkillSample]:
    return [
        SkillSample(
            prompt=str(normalize_spreadsheetbench_record(record).get("instruction", "")),
            expected_answer=str(normalize_spreadsheetbench_record(record).get("expected_answer", "")),
            metadata=normalize_spreadsheetbench_record(record),
        )
        for record in records
    ]


def extract_code(text: str) -> str:
    if "```" not in text:
        return text.strip()
    start = text.find("```")
    newline = text.find("\n", start)
    end = text.find("```", newline + 1)
    if newline == -1 or end == -1:
        return text.strip()
    return text[newline + 1 : end].strip()


def _strip_path_assignments(code: str) -> str:
    return PATH_ASSIGN_RE.sub("", code)


def run_generated_code(code: str, input_path: str, output_path: str, timeout: int = 120) -> tuple[bool, str]:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    cleaned = _strip_path_assignments(code)
    indented = textwrap.indent(cleaned, "    ")
    runner = RUNNER_TEMPLATE.format(
        input_path=input_path,
        output_path=output_path,
        user_code_indented=indented,
    )
    with tempfile.NamedTemporaryFile("w", suffix=".py", delete=False) as handle:
        handle.write(runner)
        script_path = handle.name
    try:
        process = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        if process.returncode != 0:
            return False, (process.stdout + "\n" + process.stderr).strip()
        if not os.path.exists(output_path):
            return False, "output file was not created"
        return True, ""
    except subprocess.TimeoutExpired:
        return False, f"timeout after {timeout}s"
    finally:
        try:
            os.unlink(script_path)
        except OSError:
            pass


def _datetime_to_float(value: datetime.datetime) -> float:
    excel_start = datetime.datetime(1899, 12, 30)
    delta = value - excel_start
    return delta.days + delta.seconds / 86400.0


def _transform_value(value: Any) -> Any:
    if isinstance(value, bool):
        return round(float(value), 2)
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    if isinstance(value, datetime.time):
        return str(value)[:-3]
    if isinstance(value, datetime.datetime):
        return round(_datetime_to_float(value), 0)
    if isinstance(value, str):
        try:
            return round(float(value), 2)
        except ValueError:
            return value
    return value


def _compare_cell_value(left: Any, right: Any) -> bool:
    left = _transform_value(left)
    right = _transform_value(right)
    if (left == "" and right is None) or (left is None and right == ""):
        return True
    if (left == "" and right == "") or (left is None and right is None):
        return True
    if type(left) is not type(right):
        return False
    return left == right


def _col_num2name(number: int) -> str:
    name = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _col_name2num(name: str) -> int:
    number = 0
    for char in name:
        number = number * 26 + (ord(char) - ord("A") + 1)
    return number


def _parse_range(range_str: str) -> tuple[tuple[int, int], tuple[int, int]]:
    start_cell, end_cell = range_str.split(":")
    start_col = "".join(char for char in start_cell if char.isalpha())
    start_row = "".join(char for char in start_cell if char.isdigit())
    end_col = "".join(char for char in end_cell if char.isalpha())
    end_row = "".join(char for char in end_cell if char.isdigit())
    return (_col_name2num(start_col), int(start_row)), (_col_name2num(end_col), int(end_row))


def generate_cell_names(range_str: str) -> list[str]:
    if ":" not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = _parse_range(range_str)
    cols = [_col_num2name(index) for index in range(start_col, end_col + 1)]
    return [f"{col}{row}" for col in cols for row in range(start_row, end_row + 1)]


def compare_workbooks(gold_path: str, predicted_path: str, answer_position: str) -> tuple[bool, str]:
    if not os.path.exists(predicted_path):
        return False, "file not exist"
    try:
        gold_wb = openpyxl.load_workbook(filename=gold_path, data_only=True)
        pred_wb = openpyxl.load_workbook(filename=predicted_path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return False, f"load error: {exc}"
    try:
        ok_all = True
        first_message = ""
        for spec in (answer_position or "").split(","):
            spec = spec.strip()
            if not spec:
                continue
            if "!" in spec:
                sheet_name, cell_range = spec.split("!", 1)
                sheet_name = sheet_name.strip().strip("'\"")
            else:
                sheet_name = gold_wb.sheetnames[0]
                cell_range = spec
            cell_range = cell_range.strip().strip("'\"")
            if sheet_name not in pred_wb.sheetnames:
                return False, f"worksheet not found: {sheet_name}"
            gold_ws = gold_wb[sheet_name]
            pred_ws = pred_wb[sheet_name]
            for cell_name in generate_cell_names(cell_range):
                if not _compare_cell_value(gold_ws[cell_name].value, pred_ws[cell_name].value):
                    ok_all = False
                    if not first_message:
                        first_message = (
                            f"value@{sheet_name}!{cell_name}: "
                            f"gt={gold_ws[cell_name].value!r} pred={pred_ws[cell_name].value!r}"
                        )
        return ok_all, first_message
    finally:
        gold_wb.close()
        pred_wb.close()


def _find_test_cases(task_dir: Path) -> list[tuple[str, str, str]]:
    cases: list[tuple[str, str, str]] = []
    for input_path in sorted(task_dir.glob("*_input.xlsx")):
        case_no = input_path.name.split("_", 1)[0]
        gold_path = Path(str(input_path).replace("_input.xlsx", "_answer.xlsx"))
        if gold_path.exists():
            cases.append((case_no, str(input_path), str(gold_path)))
    for input_path in sorted(task_dir.glob("*_init.xlsx")):
        case_no = input_path.name.split("_", 1)[0]
        gold_path = Path(str(input_path).replace("_init.xlsx", "_golden.xlsx"))
        if gold_path.exists():
            cases.append((case_no, str(input_path), str(gold_path)))
    if not cases:
        bare_init = task_dir / "initial.xlsx"
        bare_gold = task_dir / "golden.xlsx"
        if bare_init.exists() and bare_gold.exists():
            cases.append(("1", str(bare_init), str(bare_gold)))
    return cases


def resolve_test_cases(metadata: dict[str, Any]) -> list[tuple[str, str, str]]:
    input_path = str(metadata.get("input_path") or "").strip()
    gold_path = str(metadata.get("gold_path") or "").strip()
    if input_path and gold_path:
        return [("1", input_path, gold_path)]
    spreadsheet_path = str(metadata.get("spreadsheet_path") or "").strip()
    data_root = str(metadata.get("data_root") or "").strip()
    if not spreadsheet_path:
        return []
    task_dir = Path(spreadsheet_path)
    if not task_dir.is_absolute() and data_root:
        task_dir = Path(data_root) / spreadsheet_path
    if task_dir.is_file():
        return []
    return _find_test_cases(task_dir)


def extract_answer(text: str) -> str:
    matches = re.findall(r"<answer>(.*?)</answer>", text, re.DOTALL | re.IGNORECASE)
    if matches:
        return matches[-1].strip()
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return lines[-1] if lines else text.strip()


class SpreadsheetBenchEvaluator(SkillEvaluator):
    def evaluate(self, prediction: str, sample: SkillSample) -> MetricResult:
        predicted_answer = extract_answer(prediction)
        answer_position = str(sample.metadata.get("answer_position") or "")
        direct_gold_path = str(sample.metadata.get("gold_path") or "")
        if direct_gold_path and answer_position and os.path.exists(predicted_answer):
            ok, reason = compare_workbooks(direct_gold_path, predicted_answer, answer_position)
            return MetricResult(
                score=1.0 if ok else 0.0,
                passed=ok,
                details={
                    "ok": ok,
                    "reason": reason,
                    "predicted_answer": predicted_answer,
                    "instruction_type": sample.metadata.get("instruction_type", ""),
                    "task_type": sample.metadata.get("task_type", ""),
                },
            )

        cases = resolve_test_cases(sample.metadata)
        code = extract_code(prediction)
        if cases and answer_position and code:
            case_results: list[dict[str, Any]] = []
            passed_cases = 0
            with tempfile.TemporaryDirectory() as temp_dir:
                for case_no, input_path, gold_path in cases:
                    output_path = str(Path(temp_dir) / f"{case_no}_pred.xlsx")
                    exec_ok, exec_reason = run_generated_code(code, input_path, output_path)
                    compare_ok = False
                    compare_reason = exec_reason
                    if exec_ok:
                        compare_ok, compare_reason = compare_workbooks(gold_path, output_path, answer_position)
                    if exec_ok and compare_ok:
                        passed_cases += 1
                    case_results.append(
                        {
                            "case_no": case_no,
                            "input_path": input_path,
                            "gold_path": gold_path,
                            "output_path": output_path,
                            "exec_ok": exec_ok,
                            "ok": exec_ok and compare_ok,
                            "reason": compare_reason,
                        }
                    )
            soft = passed_cases / len(cases)
            hard = int(passed_cases == len(cases))
            return MetricResult(
                score=soft,
                passed=bool(hard),
                details={
                    "ok": bool(hard),
                    "hard": hard,
                    "soft": soft,
                    "n_cases": len(cases),
                    "n_pass": passed_cases,
                    "case_results": case_results,
                    "predicted_answer": predicted_answer,
                    "instruction_type": sample.metadata.get("instruction_type", ""),
                    "task_type": sample.metadata.get("task_type", ""),
                    "mode": "generated_code",
                },
            )

        normalized_prediction = " ".join(predicted_answer.strip().lower().split())
        normalized_expected = " ".join(str(sample.expected_answer).strip().lower().split())
        ok = normalized_prediction == normalized_expected
        return MetricResult(
            score=1.0 if ok else 0.0,
            passed=ok,
            details={
                "ok": ok,
                "reason": "" if ok else f"predicted {predicted_answer!r} but expected {sample.expected_answer!r}",
                "predicted_answer": predicted_answer,
                "instruction_type": sample.metadata.get("instruction_type", ""),
                "task_type": sample.metadata.get("task_type", ""),
            },
        )
